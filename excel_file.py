import openpyxl as xl
from openpyxl.chart import BarChart,Reference 


def file_process(filename):
    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1']
    #take values of column 3 of each row and procees it and save it in new column
    for row in range(2,sheet.max_row+1):
     cell=sheet.cell(row,3)
     corrected_price=cell.value*0.9
     corrected_price_cell=sheet.cell(row,4)
     corrected_price_cell.value=corrected_price
     

    #create a chart in excel sheet
    values=Reference(sheet,
                     min_row=2,
                     max_row=sheet.max_row,
                     min_col=4,
                     max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')


    wb.save(filename)

